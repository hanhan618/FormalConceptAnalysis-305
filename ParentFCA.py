import numpy as np
import openpyxl
import xlrd
import re

def SaveExcel(input, output):
    # Xác định số hàng và cột lớn nhất trong file excel cần tạo
    row = len(input)
    column = len(input[0])
    # Tạo một workbook mới và active
    wb = openpyxl.Workbook()
    ws = wb.active
    # ghi nội dung từ input vào file Excel
    # Dong tieu de
    ws.cell(column=1, row=1, value=input[0][0])
    ws.cell(column=2, row=1, value=input[0][1])
    ws.cell(column=3, row=1, value=input[0][2])
    ws.cell(column=4, row=1, value='Parent')
    for i in range(1, row):
        v = input[i][0]
        ws.cell(column=1, row=i + 1, value=v)
        v2 = input[i][1]
        ws.cell(column=2, row=i + 1, value=v2)
        v3 = input[i][2]
        ws.cell(column=3, row=i + 1, value=str(v3).replace('[','').replace(']','').replace("'",'').replace('  ', ''))
        t=""
        list2=[]
        list2=input[i][3]
        for f in list2:
            t+= '#' + str(f) + ', '
        ws.cell(column=4, row=i + 1, value=t)

    # Lưu lại file Excel
    wb.save(output)
# ----------------------------------
file = 'FC_covid19_patient.xlsx'

wbRead = xlrd.open_workbook(file) #workBook read
sheetRead = wbRead.sheet_by_index(0)

arr = [['ID','Extent','Intent','Parent']]

for rows in range(1,sheetRead.nrows):
    arr.append([])
    arr[rows].append(int(sheetRead.cell_value(rows, 0)))
    arr[rows].append(str(sheetRead.cell_value(rows, 1)))
    a=[]
    for ele in re.split(',', sheetRead.cell_value(rows, 2).replace('(','').replace(')','').replace("'",'')):
        if len(ele)>0:
            a.append(ele)
    arr[rows].append(a)
    arr[rows].append([0])

# print(arr)
for i in range(1, len(arr)):
    if len(arr[i][2])>1:
        s1 = {}
        s1 = arr[i][2]
        p = []
        for j in range(1, i):
            s2={}
            s2=arr[j][2]
            if set(s2).intersection(set(s1)) == set(s2):
                p.append(j)
        arr[i][3]=p
# print(arr)
# Luu file
output= 'DBCovid19_patient.xlsx'
SaveExcel(arr,output)

