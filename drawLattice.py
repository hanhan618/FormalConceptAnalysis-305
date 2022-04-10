import concepts.lattices
import openpyxl
import xlrd
from openpyxl import load_workbook
import pandas as pd
from concepts import Context
import graphviz
from graphviz import Graph

c = Context.fromfile('patient_covid19_context.csv', frmat = 'csv', encoding = 'utf-8')
# print(c)

#
# # Draw Lattice
dot = c.lattice.graphviz(view=True)
# print(dot.source)
#
# #
# # intent = c.intension(['11','17','20'])
# # print(f"Thuoc tinh chung cua benh nhan 11, 17 va 21: ", intent)
# # extent = c.extension(['ho', 'sốt', 'female'])
# # print(f"Cac doi tuong co thuoc tinh ho, sot, female: ", extent)
#
#
#
#
file = 'FC_covid19_patient.xlsx'
# # get the first sheet as an object
wbRead = xlrd.open_workbook(file) #workBook read
wbWrite = openpyxl.load_workbook(filename = file) #workBook write
sheetWriteFC = wbWrite.get_sheet_by_name("FC_location")

row = 1
for extent, intent in c.lattice:
    extent = sheetWriteFC.cell(row=row, column=1, value=str(extent).lstrip('(').rstrip(')').replace("'",''))
    intent = sheetWriteFC.cell(row=row, column=2, value=str(intent).lstrip('(').rstrip(')').replace("'",''))
    row += 1

wbWrite.save(filename = file)


# # Duyệt các Formal Concept
files = open("fc_location.txt","w", encoding="utf-8")
l = c.lattice
for index in range(1, len(l) - 1):
    files.write(str(index) + str(l[index])+ '\n')
files.close()
#
#

# for extent, intent in c.lattice:
#     print("%r %r" % (extent, intent))

